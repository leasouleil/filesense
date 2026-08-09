from openpyxl import load_workbook
from filesense.readers.base import BaseReader


class XlsxReader(BaseReader):
    def read(self, filepath: str) -> dict:
        workbook = load_workbook(filepath, data_only=True)

        text = ""
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text += str(cell) + " "
                text += "\n"

        return {"text": text, "metadata": {}}
