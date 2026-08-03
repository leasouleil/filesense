from openpyxl import load_workbook

def extract_xlsx_text(filepath):
    workbook = load_workbook(filepath, data_only=True)

    text = ""

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    text += str(cell) + " "
            text += "\n"

    return text